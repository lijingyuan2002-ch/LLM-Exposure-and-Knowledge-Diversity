import pandas as pd
import numpy as np
import warnings
import glob
from pathlib import Path
import patsy
import statsmodels.api as sm
from sklearn.preprocessing import StandardScaler
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
warnings.filterwarnings('ignore')
print("="*60)
print("Step 1: Load and Merge Datasets")
print("="*60)
df_parquet = pd.read_parquet("sciscinet_papers_filtered.parquet",
                              columns=['paperid', 'Atyp_10pct_Z', 'year'])
print(f"Raw parquet dataset shape: {df_parquet.shape}")
df_parquet = df_parquet[df_parquet['year'].isin([2023, 2024])]
df_parquet['paperid_clean'] = df_parquet['paperid'].str.replace('https://openalex.org/', '')
print("\nLoading topic mapping table...")
topic_mapping = pd.read_excel("OpenAlex_topic_mapping_table.xlsx")
topic_mapping['topic_id_with_prefix'] = 'T' + topic_mapping['topic_id'].astype(str)
csv_folder = "works_2023-2025"
csv_files = glob.glob(f"{csv_folder}/*.csv")
print(f"\nFound {len(csv_files)} CSV files")
required_columns = ['id', 'team_knowledge_variety', 'team_knowledge_distance',
                    'team_knowledge_balance', 'alpha_gt_0.1', 'authors_count',
                    'institutions_distinct_count', 'countries_distinct_count',
                    'RS', 'topic_id', 'first_author_country']
all_matches = []
for csv_file in csv_files:
    try:
        df_csv = pd.read_csv(csv_file, usecols=required_columns)
        df_csv['id_clean'] = df_csv['id'].str.replace('https://openalex.org/', '')
        merged = df_parquet.merge(df_csv, left_on='paperid_clean', right_on='id_clean', how='inner')
        if not merged.empty:
            all_matches.append(merged)
    except Exception as e:
        print(f"Error processing file: {e}")
if all_matches:
    df_final = pd.concat(all_matches, ignore_index=True)
else:
    print("No matched data found!")
    exit()
df_final = df_final.drop_duplicates(subset=['paperid_clean'])
print(f"Sample size after deduplication: {len(df_final)}")
print("\nMap topic_id to field_id...")
df_final['topic_id'] = df_final['topic_id'].astype(str)
df_final = df_final.merge(topic_mapping[['topic_id_with_prefix', 'field_id']],
                          left_on='topic_id', right_on='topic_id_with_prefix', how='left')
df_final['field_id'] = df_final['field_id'].fillna('Other').astype(str)
print("\n" + "="*60)
print("Step 2: Data Preprocessing and Cleaning")
print("="*60)
df_final['novelty'] = (df_final['Atyp_10pct_Z'] < 0).astype(int)
# Rename alpha_gt_0.1 to alpha_gt_0_1
df_final.rename(columns={'alpha_gt_0.1': 'alpha_gt_0_1'}, inplace=True)
numeric_cols = ['alpha_gt_0_1', 'team_knowledge_variety', 'team_knowledge_distance',
                'team_knowledge_balance', 'authors_count', 'institutions_distinct_count',
                'countries_distinct_count', 'RS', 'year']
for col in numeric_cols:
    df_final[col] = pd.to_numeric(df_final[col], errors='coerce')
    df_final[col] = df_final[col].replace([np.inf, -np.inf], np.nan)
df_final = df_final[df_final['authors_count'] > 0]
df_final = df_final[df_final['institutions_distinct_count'] > 0]
df_final = df_final[df_final['countries_distinct_count'] > 0]
for var in ['authors_count', 'institutions_distinct_count', 'countries_distinct_count']:
    df_final[f'log_{var}'] = np.log(df_final[var])
df_final['first_author_country'] = df_final['first_author_country'].fillna('Unknown')
analysis_vars = ['alpha_gt_0_1', 'team_knowledge_variety', 'team_knowledge_distance',
                 'team_knowledge_balance', 'novelty', 'authors_count', 'institutions_distinct_count',
                 'countries_distinct_count', 'RS', 'year', 'log_authors_count',
                 'log_institutions_distinct_count', 'log_countries_distinct_count',
                 'field_id', 'first_author_country']
df_clean = df_final[analysis_vars].dropna()
print("\nFilter countries and fields with insufficient observations...")
country_counts = df_clean['first_author_country'].value_counts()
valid_countries = country_counts[country_counts >= 30].index
df_clean = df_clean[df_clean['first_author_country'].isin(valid_countries)]
field_counts = df_clean['field_id'].value_counts()
valid_fields = field_counts[field_counts >= 30].index
df_clean = df_clean[df_clean['field_id'].isin(valid_fields)]
print(f"Final sample size after filtering: {len(df_clean)}")
print("\n" + "="*60)
print("Step 3: Standardize Continuous Variables")
print("="*60)
continuous_vars = ['team_knowledge_variety', 'team_knowledge_distance', 'team_knowledge_balance',
                   'log_authors_count', 'log_institutions_distinct_count',
                   'log_countries_distinct_count', 'RS']
df_clean_scaled = df_clean.copy()
scaler = StandardScaler()
for var in continuous_vars:
    scaled_values = scaler.fit_transform(df_clean[[var]]).flatten()
    df_clean_scaled[f'{var}_scaled'] = scaled_values
df_clean_scaled['field_id'] = df_clean_scaled['field_id'].astype('category')
df_clean_scaled['first_author_country'] = df_clean_scaled['first_author_country'].astype('category')
df_clean_scaled['year'] = df_clean_scaled['year'].astype('category')
print("\n" + "="*60)
print("Step 4: Baseline Models (Model 1-3)")
print("="*60)
y = df_clean_scaled['novelty']
# Model 1: Only LLM exposure
print("\nModel 1: LLM Exposure → Novelty")
X1 = sm.add_constant(df_clean_scaled[['alpha_gt_0_1']])
model1 = sm.Logit(y, X1).fit(disp=0)
print(f"Model 1 completed, variables: {list(model1.params.index)}")
# Model 2: Only knowledge diversity indicators
print("\nModel 2: Knowledge Indicators → Novelty")
X2 = sm.add_constant(df_clean_scaled[['team_knowledge_variety_scaled',
                                       'team_knowledge_distance_scaled',
                                       'team_knowledge_balance_scaled']])
model2 = sm.Logit(y, X2).fit(disp=0)
print(f"Model 2 completed, variables: {list(model2.params.index)}")
# Model 3: LLM exposure + knowledge indicators
print("\nModel 3: LLM Exposure + Knowledge Indicators → Novelty")
X3 = sm.add_constant(df_clean_scaled[['alpha_gt_0_1', 'team_knowledge_variety_scaled',
                                       'team_knowledge_distance_scaled', 'team_knowledge_balance_scaled']])
model3 = sm.Logit(y, X3).fit(disp=0)
print(f"Model 3 completed, variables: {list(model3.params.index)}")
print("\n" + "="*60)
print("Step 5: Full Models (Model 4-6)")
print("="*60)
control_vars_scaled = ['log_authors_count_scaled', 'log_institutions_distinct_count_scaled',
                       'log_countries_distinct_count_scaled', 'RS_scaled']
df_for_patsy = df_clean_scaled.copy()
# Model 4: LLM exposure + controls + fixed effects
print("\nModel 4: LLM Exposure + Controls + FE → Novelty")
formula4 = "novelty ~ alpha_gt_0_1 + " + " + ".join(control_vars_scaled) + " + C(field_id)"
y4, X4 = patsy.dmatrices(formula4, df_for_patsy, return_type='dataframe')
model4 = sm.Logit(y4, X4).fit(disp=0, maxiter=300)
# Model 5: Knowledge indicators + controls + fixed effects
print("\nModel 5: Knowledge Indicators + Controls + FE → Novelty")
formula5 = "novelty ~ team_knowledge_variety_scaled + team_knowledge_distance_scaled + team_knowledge_balance_scaled + " + " + ".join(control_vars_scaled) + " + C(field_id)"
y5, X5 = patsy.dmatrices(formula5, df_for_patsy, return_type='dataframe')
model5 = sm.Logit(y5, X5).fit(disp=0, maxiter=300)
# Model 6: LLM exposure + knowledge indicators + controls + fixed effects
print("\nModel 6: LLM Exposure + Knowledge Indicators + Controls + FE → Novelty")
formula6 = "novelty ~ alpha_gt_0_1 + team_knowledge_variety_scaled + team_knowledge_distance_scaled + team_knowledge_balance_scaled + " + " + ".join(control_vars_scaled) + " + C(field_id)"
y6, X6 = patsy.dmatrices(formula6, df_for_patsy, return_type='dataframe')
model6 = sm.Logit(y6, X6).fit(disp=0, maxiter=300)
print("\n" + "="*60)
print("Step 6: Export Regression Output to Excel")
print("="*60)
def format_coef_se(model, var_name):
    if var_name not in model.params.index:
        return "-\n(-)"
    coef = model.params[var_name]
    se = model.bse[var_name]
    p = model.pvalues[var_name]
    if p < 0.001:
        stars = "***"
    elif p < 0.01:
        stars = "**"
    elif p < 0.05:
        stars = "*"
    else:
        stars = ""
    return f"{coef:.5f}{stars}\n({se:.5f})"
wb = Workbook()
ws1 = wb.active
ws1.title = "Regression_Summary"
header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
cell_alignment = Alignment(horizontal='center', vertical='center')
left_alignment = Alignment(horizontal='left', vertical='center')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
model_titles = [
    '(1)\nModel1\nOnly LLM',
    '(2)\nModel2\nOnly Knowledge Indicators',
    '(3)\nModel3\nLLM + Knowledge Indicators',
    '(4)\nModel4\nLLM + Controls + FE',
    '(5)\nModel5\nKnowledge Indicators + Controls + FE',
    '(6)\nModel6\nLLM + Knowledge Indicators + Controls + FE'
]
for col, title in enumerate(model_titles, 2):
    cell = ws1.cell(row=1, column=col, value=title)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
ws1.cell(row=1, column=1, value="Variable").font = header_font
ws1.cell(row=1, column=1).fill = header_fill
ws1.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws1.cell(row=1, column=1).border = border
# Define independent variables to display
variables_order = [
    'alpha_gt_0_1',
    'team_knowledge_variety_scaled',
    'team_knowledge_distance_scaled',
    'team_knowledge_balance_scaled',
    'log_authors_count_scaled',
    'log_institutions_distinct_count_scaled',
    'log_countries_distinct_count_scaled',
    'RS_scaled',
    'Intercept'   # Constant term generated by patsy is named 'Intercept'
]
display_names = {
    'alpha_gt_0_1': 'LLM Exposure',
    'team_knowledge_variety_scaled': 'Knowledge Variety',
    'team_knowledge_distance_scaled': 'Knowledge Distance',
    'team_knowledge_balance_scaled': 'Knowledge Balance',
    'log_authors_count_scaled': 'Author Count (ln)',
    'log_institutions_distinct_count_scaled': 'Institution Count (ln)',
    'log_countries_distinct_count_scaled': 'Country Count (ln)',
    'RS_scaled': 'Research Scale RS',
    'Intercept': 'Intercept'
}
models_list = [model1, model2, model3, model4, model5, model6]
# Extract coefficients for each model
model_var_mappings = []
for model in models_list:
    m_vars = {}
    for var in variables_order:
        # Handle inconsistent constant term naming
        if var == 'Intercept':
            if 'const' in model.params.index:
                actual_var = 'const'
            elif 'Intercept' in model.params.index:
                actual_var = 'Intercept'
            else:
                actual_var = None
        else:
            actual_var = var
        if actual_var is not None:
            m_vars[var] = format_coef_se(model, actual_var)
        else:
            m_vars[var] = "-\n(-)"
    model_var_mappings.append(m_vars)
# Write data rows
for row_idx, var in enumerate(variables_order, 2):
    display_var = display_names.get(var, var)
    cell = ws1.cell(row=row_idx, column=1, value=display_var)
    cell.font = Font(bold=True)
    cell.alignment = left_alignment
    cell.border = border
    
    for col_idx, m_vars in enumerate(model_var_mappings, 2):
        value = m_vars.get(var, "-\n(-)")
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = border
        cell.font = Font(size=9)
# Fixed effects row
fe_row = len(variables_order) + 2
ws1.cell(row=fe_row, column=1, value="Fixed Effects").font = Font(bold=True)
ws1.cell(row=fe_row, column=1).alignment = left_alignment
ws1.cell(row=fe_row, column=1).border = border
fe_values = ["No Controls", "No Controls", "No Controls", "Field FE", "Field FE", "Field FE"]
for col_idx, fe_val in enumerate(fe_values, 2):
    cell = ws1.cell(row=fe_row, column=col_idx, value=fe_val)
    cell.alignment = cell_alignment
    cell.border = border
# Control variables row
control_row = fe_row + 1
ws1.cell(row=control_row, column=1, value="Control Variables").font = Font(bold=True)
ws1.cell(row=control_row, column=1).alignment = left_alignment
ws1.cell(row=control_row, column=1).border = border
control_values = ["No", "No", "No", "Yes", "Yes", "Yes"]
for col_idx, control_val in enumerate(control_values, 2):
    cell = ws1.cell(row=control_row, column=col_idx, value=control_val)
    cell.alignment = cell_alignment
    cell.border = border
# Observation count row
n_row = control_row + 1
ws1.cell(row=n_row, column=1, value="Observations").font = Font(bold=True)
ws1.cell(row=n_row, column=1).alignment = left_alignment
ws1.cell(row=n_row, column=1).border = border
for col_idx, model in enumerate(models_list, 2):
    cell = ws1.cell(row=n_row, column=col_idx, value=f"{int(model.nobs):,}")
    cell.alignment = cell_alignment
    cell.border = border
# Pseudo R-squared row
r2_row = n_row + 1
ws1.cell(row=r2_row, column=1, value="Pseudo R²").font = Font(bold=True)
ws1.cell(row=r2_row, column=1).alignment = left_alignment
ws1.cell(row=r2_row, column=1).border = border
for col_idx, model in enumerate(models_list, 2):
    cell = ws1.cell(row=r2_row, column=col_idx, value=f"{model.prsquared:.5f}")
    cell.alignment = cell_alignment
    cell.border = border
# AIC row
aic_row = r2_row + 1
ws1.cell(row=aic_row, column=1, value="AIC").font = Font(bold=True)
ws1.cell(row=aic_row, column=1).alignment = left_alignment
ws1.cell(row=aic_row, column=1).border = border
for col_idx, model in enumerate(models_list, 2):
    cell = ws1.cell(row=aic_row, column=col_idx, value=f"{model.aic:.2f}")
    cell.alignment = cell_alignment
    cell.border = border
# Adjust column width and row height
ws1.column_dimensions['A'].width = 22
for col in range(2, 8):
    ws1.column_dimensions[get_column_letter(col)].width = 24
for row in range(1, aic_row + 1):
    ws1.row_dimensions[row].height = 35 if row == 1 else 25
# Worksheet 2: Detailed regression results
ws2 = wb.create_sheet("Detailed_Regression_Results")
detail_headers = ['Model', 'Variable', 'Coefficient', 'Std_Error', 'z_value', 'p_value', 'CI95_Lower', 'CI95_Upper']
for col, header in enumerate(detail_headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
row_offset = 2
for model_idx, (model, m_name) in enumerate(zip(models_list, model_titles)):
    interesting_vars = []
    for var in variables_order:
        if var == 'Intercept':
            if 'const' in model.params.index:
                interesting_vars.append('const')
            elif 'Intercept' in model.params.index:
                interesting_vars.append('Intercept')
        else:
            if var in model.params.index:
                interesting_vars.append(var)
    for var_idx, var_name in enumerate(interesting_vars):
        conf_int = model.conf_int().loc[var_name] if var_name in model.conf_int().index else (None, None)
        row = row_offset + var_idx
        ws2.cell(row=row, column=1, value=m_name).border = border
        ws2.cell(row=row, column=2, value=var_name).border = border
        ws2.cell(row=row, column=3, value=f"{model.params[var_name]:.6f}").border = border
        ws2.cell(row=row, column=4, value=f"{model.bse[var_name]:.6f}").border = border
        ws2.cell(row=row, column=5, value=f"{model.tvalues[var_name]:.4f}").border = border
        ws2.cell(row=row, column=6, value=f"{model.pvalues[var_name]:.6f}").border = border
        ws2.cell(row=row, column=7, value=f"{conf_int[0]:.6f}" if conf_int[0] else 'N/A').border = border
        ws2.cell(row=row, column=8, value=f"{conf_int[1]:.6f}" if conf_int[1] else 'N/A').border = border
        for col in range(1, 9):
            ws2.cell(row=row, column=col).alignment = cell_alignment
    row_offset += len(interesting_vars) + 2
for col in range(1, 9):
    ws2.column_dimensions[get_column_letter(col)].width = 18
output_file = "logit_regression_results.xlsx"
os.makedirs(os.path.dirname(output_file), exist_ok=True)
wb.save(output_file)
print(f"\n✅ Regression results successfully saved to: {output_file}")